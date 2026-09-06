# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""XLSX assembler for the Certification Test Case Document.

The `cert_test_cases` Product Kit output is generated as markdown with one or
more pipe-delimited tables. The Authority certification teams need the same content
as a .xlsx workbook (matching the layout in past cert testcase XLSX files
that we ingested under DocCategory.CERT_TESTCASE).

This module parses the markdown tables and emits an openpyxl workbook with:
  - One sheet per section heading (### 4.1 Happy Path ... → "Happy Path")
  - Styled header row (bold, filled)
  - Auto-widened columns
  - [NEEDS_PM_INPUT] cells highlighted in yellow

Callers consume `assemble_cert_testcase_xlsx(markdown) -> bytes` and stream
the bytes back via an API endpoint.
"""
from __future__ import annotations

import io
import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_NEEDS_INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")

_SHEET_NAME_MAX = 31  # Excel cap


def _clean_sheet_name(raw: str) -> str:
    """Normalise a section heading to a valid Excel sheet name (<=31 chars, no /\\?*[])."""
    name = re.sub(r"[\\/:*?\[\]]", "", raw).strip()
    if not name:
        name = "Test Cases"
    return name[:_SHEET_NAME_MAX]


def _parse_markdown_tables(md: str) -> list[tuple[str, list[str], list[list[str]]]]:
    """Extract pipe-delimited tables grouped by their preceding ### heading.

    Returns list of (sheet_name, header_row, data_rows). If no ### heading is
    found above a table, the sheet is named "Test Cases".
    """
    lines = md.splitlines()

    sections: list[tuple[str, list[str], list[list[str]]]] = []
    current_heading = "Test Cases"
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        heading_match = re.match(r"^#{2,4}\s+(.+?)\s*$", line)
        if heading_match:
            current_heading = heading_match.group(1).strip()
            i += 1
            continue

        # Table header row detection: a pipe-line followed by a separator line of | --- | --- |
        if "|" in line and i + 1 < len(lines):
            sep = lines[i + 1].strip()
            if re.match(r"^\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$", sep):
                header_cells = [c.strip() for c in _split_md_row(line)]
                data_rows: list[list[str]] = []
                j = i + 2
                while j < len(lines) and "|" in lines[j].strip():
                    row_line = lines[j].strip()
                    # Stop at blank line masquerading as pipe row
                    if not row_line.replace("|", "").strip():
                        j += 1
                        continue
                    row_cells = [c.strip() for c in _split_md_row(row_line)]
                    # Pad / trim to header width
                    if len(row_cells) < len(header_cells):
                        row_cells += [""] * (len(header_cells) - len(row_cells))
                    elif len(row_cells) > len(header_cells):
                        row_cells = row_cells[: len(header_cells)]
                    data_rows.append(row_cells)
                    j += 1
                if header_cells and data_rows:
                    sections.append((_clean_sheet_name(current_heading), header_cells, data_rows))
                i = j
                continue

        i += 1

    return sections


def _split_md_row(row: str) -> list[str]:
    """Split a markdown table row into cells, stripping leading/trailing pipes."""
    stripped = row.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return stripped.split("|")


def assemble_cert_testcase_xlsx(markdown: str, feature_name: str = "the network Feature") -> bytes:
    """Convert a cert-testcase markdown document into an XLSX workbook.

    Returns the bytes of the .xlsx file. Never raises on malformed markdown —
    falls back to a single sheet with the raw markdown embedded if no tables
    can be parsed.
    """
    sections = _parse_markdown_tables(markdown or "")

    wb = Workbook()
    # Drop the default blank sheet — we'll add our own
    default_ws = wb.active
    wb.remove(default_ws)

    if not sections:
        logger.warning("xlsx_assembler: no tables found, emitting fallback sheet")
        ws = wb.create_sheet("Test Cases")
        ws.cell(row=1, column=1, value=f"{feature_name} — Certification Test Cases").font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="(Could not parse tables from generated markdown. Raw output below:)")
        for idx, line in enumerate((markdown or "").splitlines()[:1000], start=5):
            ws.cell(row=idx, column=1, value=line)
        ws.column_dimensions["A"].width = 120
    else:
        used_names: set[str] = set()
        for raw_name, header, rows in sections:
            # Dedup sheet names ("Happy Path" clash etc.)
            name = raw_name
            n = 2
            while name in used_names:
                suffix = f" ({n})"
                name = (raw_name[: _SHEET_NAME_MAX - len(suffix)]) + suffix
                n += 1
            used_names.add(name)

            ws = wb.create_sheet(name)

            # Header row
            for c_idx, header_cell in enumerate(header, start=1):
                cell = ws.cell(row=1, column=c_idx, value=header_cell)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _WRAP

            # Data rows
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = _WRAP
                    if val and "[NEEDS_PM_INPUT]" in val:
                        cell.fill = _NEEDS_INPUT_FILL

            # Auto-ish column widths — capped to keep the sheet usable
            for c_idx, header_cell in enumerate(header, start=1):
                max_len = len(header_cell)
                for row in rows:
                    val = row[c_idx - 1] if c_idx - 1 < len(row) else ""
                    max_len = max(max_len, min(len(val or ""), 120))
                ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 60)

            ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    logger.info(
        "xlsx_assembler: produced %d bytes across %d sheet(s)",
        len(data), max(1, len(sections)),
    )
    return data

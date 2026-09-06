# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Programmatic repair helpers for common Excel formula defects."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .recalc_helper import CellError


def fix_formula_errors(path: Path, errors: list[CellError]) -> bool:
    """Remove obviously broken formulas that cannot be resolved locally."""

    if not errors:
        return False
    wb = load_workbook(path)
    changed = False
    for error in errors:
        ws = wb[error.sheet]
        cell = ws[error.cell]
        if isinstance(cell.value, str) and "#REF!" in cell.value:
            cell.value = ""
            changed = True
    if changed:
        wb.save(path)
    return changed

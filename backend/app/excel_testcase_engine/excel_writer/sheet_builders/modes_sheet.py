# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Build MODES OF CERTIFICATION matrix."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from app.excel_testcase_engine.excel_writer.styles import ALIGN_BODY_CENTER, BORDER, FONT_BODY, FONT_HEADER, HEADER_FILL, TAB_COLORS
from app.excel_testcase_engine.schemas.workbook_plan import WorkbookPlan


def build(ws: Worksheet, plan: WorkbookPlan) -> None:
    testcase_sheets = [sheet for sheet in plan.sheets if sheet.test_cases]
    for idx, sheet in enumerate(testcase_sheets, start=1):
        ws.cell(3, idx, f"Subset-{idx}")
        ws.cell(4, idx, sheet.name)
        for row_offset, tc in enumerate(sheet.test_cases, start=5):
            ws.cell(row_offset, idx, tc.test_id)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = FONT_HEADER if cell.row in {3, 4} else FONT_BODY
            cell.border = BORDER
            cell.alignment = ALIGN_BODY_CENTER
            if cell.row in {3, 4}:
                cell.fill = HEADER_FILL
    for idx in range(1, max(len(testcase_sheets), 1) + 1):
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = 22
    ws.sheet_properties.tabColor = TAB_COLORS["meta"]

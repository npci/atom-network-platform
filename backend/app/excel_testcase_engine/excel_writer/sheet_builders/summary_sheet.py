# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Build Archetype C summary sheet."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from app.excel_testcase_engine.excel_writer.styles import ALIGN_BODY_CENTER, ALIGN_BODY_LEFT, BORDER, FONT_BODY, FONT_HEADER, HEADER_FILL, TAB_COLORS
from app.excel_testcase_engine.schemas.workbook_plan import WorkbookPlan


def build(ws: Worksheet, plan: WorkbookPlan) -> None:
    from openpyxl.utils import get_column_letter

    from app.excel_testcase_engine import domain_vocab

    # Role columns and scope names come from the active pack (UPI: Remitter /
    # Beneficiary / Payer / Payee with Acquirer+Issuer scopes — byte-identical
    # to the previous hardcoded layout); formulas are derived from the actual
    # column positions so a domain with fewer roles still sums correctly.
    roles = domain_vocab.summary_role_headers()
    scopes = domain_vocab.scope_titles()
    extras = [f"Extra-{sc}" for sc in scopes]
    ws["A1"] = "SUMMARY SHEET"
    ws["A1"].font = FONT_HEADER
    ws.append([])
    ws.append(["Type", "Meta", *roles, "Total Mandatory cases", "Mobile App Testing", *extras, "Total"])
    total_cases = sum(len(sheet.test_cases) for sheet in plan.sheets)
    mand_end = 2 + len(roles)                       # Meta + role columns
    total_mand_col = mand_end + 1
    grand_end = total_mand_col + 1 + len(extras)    # + Mobile App Testing + extras
    role_cells = [0] * len(roles)
    if role_cells:
        role_cells[max(0, len(role_cells) - 2)] = total_cases
    ws.append(["+".join(scopes) or "All roles", 0, *role_cells,
               f"=SUM(B4:{get_column_letter(mand_end)}4)", 0, *([0] * len(extras)),
               f"=SUM({get_column_letter(total_mand_col)}4:{get_column_letter(grand_end)}4)"])
    start = 7
    ws.cell(start, 1, "Sample Proposition")
    samples = [
        "1. If a new bank is onboarding as Issuer, execute issuer scoped cases.",
        "2. If a new bank is onboarding as Acquirer, execute acquirer scoped cases.",
        "3. If both roles are in scope, execute all mandatory cases.",
        "4. Execute extras when the corresponding optional feature is enabled.",
        "5. Maintain transaction IDs during certification execution.",
        f"6. Share completed evidence with {domain_vocab.authority_label()} certification reviewers.",
    ]
    for idx, line in enumerate(samples, start=start + 1):
        ws.cell(idx, 1, line)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = FONT_HEADER if cell.row in {1, 3, start} else FONT_BODY
            cell.border = BORDER
            cell.alignment = ALIGN_BODY_LEFT if cell.column == 1 else ALIGN_BODY_CENTER
            if cell.row == 3:
                cell.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 40
    for col_idx in range(2, grand_end + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.sheet_properties.tabColor = TAB_COLORS["meta"]

# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Build formatted the Authority test-case sheets.

Quality features:
- HOME hyperlink for Archetype C role sheets.
- Header row formatted with HEADER_FILL + bold + border + 30px row.
- Body rows wrap-text, 140px row height, full border.
- ``tc.highlight=True`` paints the row yellow (HIGHLIGHT_FILL).
- STATUS column gets a data-validation dropdown of accepted values.
- Page setup primed for landscape A4 with fit-to-page so packs print clean.
- Print titles freeze the header so it repeats on every printed page.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.excel_testcase_engine.excel_writer.layouts import LAYOUT_REGISTRY
from app.excel_testcase_engine.excel_writer.styles import (
    ALIGN_BODY_CENTER,
    ALIGN_BODY_LEFT,
    ALIGN_HEADER,
    BORDER,
    FONT_BODY,
    FONT_HEADER,
    FONT_LINK,
    HEADER_FILL,
    HIGHLIGHT_FILL,
    TAB_COLORS,
)
from app.excel_testcase_engine.schemas.workbook_plan import SheetSpec, TestCaseStub

_STATUS_OPTIONS_UPPER = '"SUCCESS,FAILURE,DEEMED,TIMEOUT,PARTIAL"'
_STATUS_OPTIONS_TITLE = '"Success,Failure,Deemed,Timeout,Partial,Success/partial,Pending,Passed"'


def _display_scope(tc: TestCaseStub) -> str:
    """The "Scope" cell — the case's own value, else the active pack's label.

    The stub default is empty precisely so this can resolve per-domain; the old
    hardcoded "UPI 2.0" default printed a payments version number on library-loan
    workbooks.
    """
    from app.core.domain.registry import prompt_block
    return (tc.scope or "").strip() or prompt_block("test_case_scope", "")


def _display_initiator(tc: TestCaseStub) -> str:
    """The "TXN INITIATED BY" cell — the pack's NAME for the canonical wire token.

    `tc.txn_initiated_by` stays "Bank" | "NPCI" because cert-agent matches those
    exactly (see the field's note in workbook_plan). Only the rendered label is
    per-domain, so an NLLN workbook reads "NLLC" where a UPI one reads "NPCI"
    while both push the identical value on the wire.
    """
    from app.core.domain.registry import prompt_block
    raw = (tc.txn_initiated_by or "").strip()
    if raw.upper() == "NPCI":
        return prompt_block("initiator_authority_label", raw)
    if raw.upper() == "BANK":
        return prompt_block("initiator_participant_label", raw)
    return raw


def _rendered_value(tc: TestCaseStub, kind: str) -> str:
    rendered = tc.rendered
    if rendered is None:
        if kind == "details":
            return f"API Involved: {', '.join(tc.apis)}\nType : {tc.api_type}\nEntity Involved: {', '.join(tc.entities)}\n"
        if kind == "description":
            return f"To verify {tc.scenario_summary}."
        return f"1. Execute {', '.join(tc.apis)}.\n2. Validate result - {tc.expected_status}."
    return getattr(rendered, f"{kind}_block")


def _row_for_layout(layout_key: str, index: int, tc: TestCaseStub) -> list[object]:
    details = _rendered_value(tc, "details")
    description = _rendered_value(tc, "description")
    steps = _rendered_value(tc, "steps")
    if layout_key == "A1":
        return [tc.test_id, details, description, steps, tc.expected_status.upper()]
    if layout_key == "B1":
        return [tc.test_id, details, description, tc.expected_status.upper(), steps]
    if layout_key == "C1":
        return [index, tc.test_id, details, description, _display_scope(tc),
                _display_initiator(tc), tc.expected_status, steps]
    if layout_key == "C2":
        return [index, tc.test_id, details, tc.psp_as, description, _display_scope(tc),
                _display_initiator(tc), tc.expected_status, steps]
    if layout_key == "C3":
        return [index, tc.test_id, details, description, steps, tc.expected_status, tc.entities[0] if tc.entities else "", "", ""]
    raise ValueError(f"Unsupported layout {layout_key}")


def _status_col_index(layout_key: str, start_col: int) -> int:
    """Return the 1-indexed STATUS column for a given layout."""

    return {
        "A1": start_col + 4,
        "B1": start_col + 3,
        "C1": start_col + 6,
        "C2": start_col + 7,
        "C3": start_col + 5,
    }[layout_key]


def build(ws: Worksheet, sheet: SheetSpec, archetype: str = "A") -> None:
    """Write a formatted test-case sheet."""

    layout = LAYOUT_REGISTRY[sheet.layout]

    if archetype == "C" and sheet.layout in {"C1", "C2"}:
        ws["B1"] = "HOME"
        ws["B1"].hyperlink = "#'Index'!A1"
        ws["B1"].font = FONT_LINK

    for offset, header in enumerate(layout.headers):
        cell = ws.cell(row=layout.start_row, column=layout.start_col + offset, value=header)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER
        letter = get_column_letter(layout.start_col + offset)
        ws.column_dimensions[letter].width = layout.widths.get(header, 18)
    ws.row_dimensions[layout.start_row].height = 30

    for index, tc in enumerate(sheet.test_cases, start=1):
        row = layout.start_row + index
        values = _row_for_layout(sheet.layout, index, tc)
        is_failure = (tc.expected_status or "").lower() in {"failure"}
        for offset, value in enumerate(values):
            cell = ws.cell(row=row, column=layout.start_col + offset, value=value)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = ALIGN_BODY_LEFT if isinstance(value, str) and "\n" in value else ALIGN_BODY_CENTER
            if tc.highlight:
                cell.fill = HIGHLIGHT_FILL
        ws.row_dimensions[row].height = 140

    ws.freeze_panes = layout.freeze_panes
    last_row = max(layout.start_row + len(sheet.test_cases), layout.start_row)
    last_col = layout.start_col + len(layout.headers) - 1
    ws.auto_filter.ref = f"{ws.cell(layout.start_row, layout.start_col).coordinate}:{ws.cell(last_row, last_col).coordinate}"
    ws.sheet_properties.tabColor = sheet.tab_color or TAB_COLORS["testcase"]

    # STATUS column data-validation dropdown
    if sheet.test_cases:
        casing = "upper" if archetype != "C" else "title"
        formula = _STATUS_OPTIONS_UPPER if casing == "upper" else _STATUS_OPTIONS_TITLE
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Use a canonical STATUS value."
        validation.errorTitle = "Invalid STATUS"
        validation.prompt = "Select a status value."
        validation.promptTitle = "STATUS"
        ws.add_data_validation(validation)
        status_col = _status_col_index(sheet.layout, layout.start_col)
        validation.add(
            f"{get_column_letter(status_col)}{layout.start_row + 1}:"
            f"{get_column_letter(status_col)}{last_row}"
        )

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{layout.start_row}:{layout.start_row}"
    ws.print_options.horizontalCentered = True
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4

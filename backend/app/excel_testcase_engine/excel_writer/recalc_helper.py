# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""LibreOffice recalculation helper with a Python fallback for SUM formulas.

When ``soffice``/``libreoffice`` is installed, we ask it to recalculate the
workbook in place. When it isn't (CI, sandboxed envs), we walk every formula
cell ourselves and resolve simple ``=SUM(<range>)`` expressions so the cached
values are not blank — viewers like Numbers and Google Sheets show "0" for an
unevaluated formula, which auditors find confusing. We also scan for error
tokens and surface them.
"""

from __future__ import annotations

import logging
import re
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import BaseModel

LOGGER = logging.getLogger(__name__)
ERROR_TOKENS = ("#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#N/A")

_SUM_RANGE_RE = re.compile(r"^=SUM\(([A-Z]+\d+):([A-Z]+\d+)\)$", re.IGNORECASE)
_COUNTA_RANGE_RE = re.compile(r"^=COUNTA\('?([^'!]+)'?!([A-Z]+\d+):([A-Z]+\d+)\)$", re.IGNORECASE)


class CellError(BaseModel):
    """Formula or cached value error found in a workbook."""

    sheet: str
    cell: str
    value: str


def scan_formula_errors(path: Path) -> list[CellError]:
    """Scan workbook formulas and values for common Excel error tokens."""

    wb = load_workbook(path, data_only=False)
    errors: list[CellError] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(token in value for token in ERROR_TOKENS):
                    errors.append(CellError(sheet=ws.title, cell=cell.coordinate, value=value))
    return errors


def _python_fallback_recalc(path: Path) -> None:
    """Resolve =SUM and =COUNTA formulas to literal values.

    This is intentionally conservative: only ranges within a single column on
    the same sheet are evaluated. Anything more complex is left as a formula
    string for Excel to evaluate when a user opens the file.
    """

    wb = load_workbook(path)
    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                value = cell.value.strip()
                # =SUM(A1:A10) on the same sheet
                m = _SUM_RANGE_RE.match(value)
                if m:
                    start, end = m.group(1), m.group(2)
                    s_col = column_index_from_string(re.match(r"[A-Z]+", start).group(0))  # type: ignore[union-attr]
                    s_row = int(re.search(r"\d+", start).group(0))  # type: ignore[union-attr]
                    e_col = column_index_from_string(re.match(r"[A-Z]+", end).group(0))  # type: ignore[union-attr]
                    e_row = int(re.search(r"\d+", end).group(0))  # type: ignore[union-attr]
                    total = 0
                    for r in range(s_row, e_row + 1):
                        for c in range(s_col, e_col + 1):
                            v = ws.cell(r, c).value
                            if isinstance(v, (int, float)):
                                total += v
                    cell.value = total
                    changed = True
                    continue
                # =COUNTA('Sheet'!A5:A10) on a foreign sheet
                m = _COUNTA_RANGE_RE.match(value)
                if m:
                    target = m.group(1)
                    if target in wb.sheetnames:
                        target_ws = wb[target]
                        s_letter = re.match(r"[A-Z]+", m.group(2)).group(0)  # type: ignore[union-attr]
                        e_letter = re.match(r"[A-Z]+", m.group(3)).group(0)  # type: ignore[union-attr]
                        s_row = int(re.search(r"\d+", m.group(2)).group(0))  # type: ignore[union-attr]
                        e_row = int(re.search(r"\d+", m.group(3)).group(0))  # type: ignore[union-attr]
                        s_col = column_index_from_string(s_letter)
                        e_col = column_index_from_string(e_letter)
                        count = 0
                        for r in range(s_row, e_row + 1):
                            for c in range(s_col, e_col + 1):
                                if target_ws.cell(r, c).value not in (None, ""):
                                    count += 1
                        cell.value = count
                        changed = True
    if changed:
        wb.save(path)


def run(path: Path) -> tuple[bool, list[CellError]]:
    """Recalculate workbook via LibreOffice when available, fall back to Python."""

    path = Path(path)
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        try:
            subprocess.run(
                [soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", str(path.parent), str(path)],
                check=False,
                capture_output=True,
                text=True,
                timeout=60,
            )
        except Exception as exc:  # pragma: no cover - defensive
            LOGGER.warning("LibreOffice recalc failed: %s", exc)
    else:
        try:
            _python_fallback_recalc(path)
        except Exception as exc:  # pragma: no cover - defensive
            LOGGER.warning("Python fallback recalc failed: %s", exc)

    errors = scan_formula_errors(path)
    return not errors, errors

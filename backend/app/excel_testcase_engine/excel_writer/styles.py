# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Openpyxl style constants for the Authority-format workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
HIGHLIGHT_FILL = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")

THIN = Side(style="thin", color="000000")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FONT_HEADER = Font(name="Calibri", size=11, bold=True)
FONT_BODY = Font(name="Calibri", size=11)
FONT_LINK = Font(name="Calibri", size=11, bold=True, color="0563C1", underline="single")

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_BODY_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_BODY_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

TAB_COLORS = {
    "meta": "A6A6A6",
    "testcase": "9DC3E6",
    "extras": "F4B183",
    "mandate": "B4A7D6",
}

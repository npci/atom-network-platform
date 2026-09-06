# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Render WorkbookPlan models into deterministic openpyxl workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.excel_testcase_engine.excel_writer.sheet_builders import (
    index_sheet,
    modes_sheet,
    scope_sheet,
    subset_sheet,
    summary_sheet,
    testcase_sheet,
    uat_mobile,
    validation_report,
    version_log,
)
from app.excel_testcase_engine.schemas.validation_report import CellPatch, Defect
from app.excel_testcase_engine.schemas.workbook_plan import WorkbookPlan

from . import recalc_helper, repair


def _safe_title(name: str) -> str:
    return name[:31]


def render(plan: WorkbookPlan, output_path: Path) -> Path:
    """Render a workbook plan to an .xlsx file."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    if plan.archetype == "C":
        index_ws = wb.create_sheet("Index")
        index_sheet.build(index_ws, plan)
        summary_sheet.build(wb.create_sheet("SUMMARY SHEET"), plan)
        subset_sheet.build(wb.create_sheet("SUBSET"), plan)
        modes_sheet.build(wb.create_sheet("MODES OF CERTIFICATION"), plan)
    elif plan.archetype == "B":
        version_log.build(wb.create_sheet("Version Log"), plan)

    for sheet in plan.sheets:
        if sheet.layout == "scope":
            scope_sheet.build(wb.create_sheet(_safe_title(sheet.name)), plan)
        elif sheet.layout == "uat_mobile":
            uat_mobile.build(wb.create_sheet(_safe_title(sheet.name)))
        elif sheet.layout in {"A1", "B1", "C1", "C2", "C3"}:
            testcase_sheet.build(wb.create_sheet(_safe_title(sheet.name)), sheet, plan.archetype)

    wb.save(output_path)
    for _ in range(2):
        ok, errors = recalc_helper.run(output_path)
        if ok:
            break
        if not repair.fix_formula_errors(output_path, errors):
            break
    return output_path


def apply_patches(path: Path, patches: list[CellPatch]) -> Path:
    """Apply cell patches to an existing workbook."""

    wb = load_workbook(path)
    for patch in patches:
        if patch.sheet in wb.sheetnames:
            wb[patch.sheet][f"{patch.column}{patch.row}"] = patch.new_value
    wb.save(path)
    return path


def append_validation_report_sheet(path: Path, defects: list[Defect]) -> Path:
    """Append the fall-through validation report sheet."""

    wb = load_workbook(path)
    if "Validation_Report" in wb.sheetnames:
        del wb["Validation_Report"]
    validation_report.build(wb.create_sheet("Validation_Report"), defects)
    wb.save(path)
    return path

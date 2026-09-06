# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Repairer agent: produce CellPatch list to address critical defects.

LLM-driven. After the LLM call, we additionally run a small mechanical pass
that fixes ``pair_drift`` (copy canonical sibling's DETAILS+DESCRIPTION) and
``missing_failure_code`` (append the canonical error-code clause from the
stub's ``response_code``). These mechanical fixes are NOT test-case content
generation — they patch existing rows by copying or templating from data the
plan already carries — so we keep them as a safety net even though the rest
of the pipeline is LLM-only.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from app.excel_testcase_engine.excel_writer.layouts import LAYOUT_REGISTRY
from app.excel_testcase_engine.adapters.llm import get_client
from app.excel_testcase_engine.schemas.llm import Message, SystemBlock
from app.excel_testcase_engine.observability import get_logger
from app.excel_testcase_engine.schemas.validation_report import CellPatch, Defect
from app.excel_testcase_engine.schemas.workbook_plan import WorkbookPlan

from ._runtime import load_prompt, parse_json_response

LOGGER = get_logger("network.agent.repairer")


def _column_letter_for(plan: WorkbookPlan, sheet_name: str, slot: str) -> str | None:
    """Return the Excel column letter for a layout slot (`details`, `description`, `steps`)."""

    sheet = next((s for s in plan.sheets if s.name == sheet_name), None)
    if sheet is None or sheet.layout not in LAYOUT_REGISTRY:
        return None
    layout = LAYOUT_REGISTRY[sheet.layout]
    if sheet.layout == "A1":
        offsets = {"details": 1, "description": 2, "steps": 3, "status": 4}
    elif sheet.layout == "B1":
        offsets = {"details": 1, "description": 2, "status": 3, "steps": 4}
    elif sheet.layout == "C1":
        offsets = {"details": 2, "description": 3, "steps": 7}
    elif sheet.layout == "C2":
        offsets = {"details": 2, "description": 4, "steps": 8}
    elif sheet.layout == "C3":
        offsets = {"details": 2, "description": 3, "steps": 4}
    else:
        return None
    if slot not in offsets:
        return None
    return get_column_letter(layout.start_col + offsets[slot])


def _rule_based_repair(plan: WorkbookPlan, path: Path, defects: list[Defect]) -> list[CellPatch]:
    """Fix pair_drift and missing_failure_code without an LLM."""

    patches: list[CellPatch] = []
    wb = load_workbook(path, data_only=False)

    # Group pair_drift defects by pair_id-via-sheet, find the canonical row, copy DETAILS+DESC.
    pair_groups: dict[str, list[Defect]] = defaultdict(list)
    for defect in defects:
        if defect.type == "pair_drift":
            pair_groups[f"{defect.sheet}|{defect.test_id or defect.row}"].append(defect)

    for sheet in plan.sheets:
        if sheet.layout not in LAYOUT_REGISTRY:
            continue
        layout = LAYOUT_REGISTRY[sheet.layout]
        title = sheet.name[:31]
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        rendered_by_id = {tc.test_id: tc for tc in sheet.test_cases}
        # find canonical (first) member per pair_id
        canonical_row_by_pair: dict[str, int] = {}
        for offset, tc in enumerate(sheet.test_cases):
            if not tc.pair_id:
                continue
            row = layout.start_row + 1 + offset
            canonical_row_by_pair.setdefault(tc.pair_id, row)
        for offset, tc in enumerate(sheet.test_cases):
            if not tc.pair_id or tc.pair_id not in canonical_row_by_pair:
                continue
            row = layout.start_row + 1 + offset
            canonical_row = canonical_row_by_pair[tc.pair_id]
            if row == canonical_row:
                continue
            details_col = _column_letter_for(plan, sheet.name, "details")
            desc_col = _column_letter_for(plan, sheet.name, "description")
            if not details_col or not desc_col:
                continue
            canonical_details = ws[f"{details_col}{canonical_row}"].value
            canonical_desc = ws[f"{desc_col}{canonical_row}"].value
            current_details = ws[f"{details_col}{row}"].value
            current_desc = ws[f"{desc_col}{row}"].value
            if current_details != canonical_details:
                patches.append(CellPatch(sheet=title, row=row, column=details_col, new_value=str(canonical_details or "")))
            if current_desc != canonical_desc:
                patches.append(CellPatch(sheet=title, row=row, column=desc_col, new_value=str(canonical_desc or "")))

    # Missing failure code: append the canonical code from the stub.
    for defect in defects:
        if defect.type != "missing_failure_code":
            continue
        sheet = next((s for s in plan.sheets if s.name == defect.sheet), None)
        if sheet is None or sheet.layout not in LAYOUT_REGISTRY:
            continue
        steps_col = _column_letter_for(plan, sheet.name, "steps")
        if not steps_col or not defect.row:
            continue
        layout = LAYOUT_REGISTRY[sheet.layout]
        offset = defect.row - layout.start_row - 1
        if offset < 0 or offset >= len(sheet.test_cases):
            continue
        tc = sheet.test_cases[offset]
        if not tc.response_code or tc.response_code == "00":
            continue
        title = sheet.name[:31]
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        current = ws[f"{steps_col}{defect.row}"].value or ""
        if not isinstance(current, str):
            continue
        patched = current.rstrip()
        if not patched.endswith("."):
            patched += "."
        from app.excel_testcase_engine import domain_vocab
        resp_api = (next((a for a in tc.apis if a.startswith("Resp")), "")
                    or domain_vocab.default_response_api() or "The response")
        patched += f'\n{len(current.splitlines()) + 1}. {resp_api} is returned with error code "{tc.response_code}".'
        patches.append(CellPatch(sheet=title, row=defect.row, column=steps_col, new_value=patched))

    return patches


async def repair(plan: WorkbookPlan, path: Path, defects: list[Defect]) -> list[CellPatch]:
    """Return cell patches that fix critical defects.

    Strategy: ask the LLM for prose-aware patches first, then fold in any
    mechanical patches (pair_drift copy, missing_failure_code append) that the
    LLM didn't address. Mechanical patches don't create test-case content —
    they reuse what the plan already carries.
    """

    critical = [d for d in defects if d.severity == "critical"]
    if not critical:
        return []

    client = get_client("repairer")

    # Provide row contents for every defect to give the LLM enough context.
    wb = load_workbook(path, data_only=False)
    enriched_defects = []
    for defect in critical:
        if defect.row and defect.sheet[:31] in wb.sheetnames:
            ws = wb[defect.sheet[:31]]
            row_values = [c.value for c in ws[defect.row]]
        else:
            row_values = []
        enriched_defects.append({**defect.model_dump(), "row_values": row_values})

    system = [SystemBlock(text=load_prompt("repairer.md"), cache=True)]
    user_msg = (
        "Workbook archetype: " + plan.archetype + "\n"
        "Defects to repair (with row context):\n"
        f"{enriched_defects}\n\n"
        "Return a JSON object with key `patches` (list of CellPatch)."
    )
    llm_patches: list[CellPatch] = []
    try:
        response = await client.complete(
            system=system,
            messages=[Message(role="user", content=user_msg)],
            max_tokens=4000,
            response_format="json",
        )
        payload = parse_json_response(response.text)
        if isinstance(payload, list):
            payload = {"patches": payload}
        llm_patches = [CellPatch.model_validate(p) for p in payload.get("patches", [])]
        LOGGER.info("repairer.llm_ok", patches=len(llm_patches), defects=len(critical))
    except (ValidationError, ValueError) as exc:
        LOGGER.warning("repairer.invalid_json", error=repr(exc)[:200])
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("repairer.provider_error", error=repr(exc))

    # Mechanical safety net: anything the LLM didn't address gets a deterministic patch.
    addressed_cells = {(p.sheet, p.row, p.column) for p in llm_patches}
    mechanical = [p for p in _rule_based_repair(plan, path, critical) if (p.sheet, p.row, p.column) not in addressed_cells]
    LOGGER.info("repairer.combined", llm=len(llm_patches), mechanical=len(mechanical))
    return llm_patches + mechanical

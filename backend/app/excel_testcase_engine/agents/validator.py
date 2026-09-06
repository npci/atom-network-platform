# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

"""Workbook validator: mechanical (Python) + semantic (LLM, parallel per sheet).

Mechanical checks (no LLM):
1.  formula errors (#REF!/#NAME?/etc.)
2.  empty TEST ID rows
3.  duplicate TEST IDs within a sheet
4.  prefix mismatches
5.  status enum violations
6.  header text mismatches
7.  wrap_text=False on multi-line cells
8.  HOME hyperlink broken (Archetype C)
9.  coverage drift (rendered vs plan.coverage_audit)
10. test ID gaps
11. merged ranges intersecting data rows
12. hardcoded numbers where formulas required (Total rows)
13. pair drift (same pair_id, different details/description)

BRD/TSD-only refactor: XSD-field validation, scope-ownership, coverage-minima,
error-code-in-scope, FR-link, and canonical-tag checks are all gone. The
allowlist for the step-lint promotion comes from the TSD interface_spec plus
whatever the stubs declare, not from a bundled UPI catalog.
"""

from __future__ import annotations

import asyncio
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from app.excel_testcase_engine import domain_vocab
from app.excel_testcase_engine.config import load_runtime_config
from app.excel_testcase_engine.excel_writer.layouts import LAYOUT_REGISTRY
from app.excel_testcase_engine.excel_writer.recalc_helper import scan_formula_errors
from app.excel_testcase_engine.adapters.llm import get_client
from app.excel_testcase_engine.schemas.llm import Message, SystemBlock
from app.excel_testcase_engine.observability import get_logger
from app.excel_testcase_engine.schemas.validation_report import Defect, ValidationReport
from app.excel_testcase_engine.schemas.workbook_plan import SheetSpec, WorkbookPlan

from ._runtime import load_prompt, parse_json_response
from .step_linter import lint_plan

LOGGER = get_logger("network.agent.validator")

_VALID_STATUSES = {"Success", "Failure", "Deemed", "Timeout", "Partial", "Success/partial", "Pending", "Passed",
                   "SUCCESS", "FAILURE", "DEEMED", "TIMEOUT", "PARTIAL"}
_TESTID_RE = re.compile(r"^[A-Z][A-Z_0-9]*_(\d+)$")

_ERROR_CODE_RE = re.compile(r"\b(?:[A-Z][0-9A-Z]{1,3}|[0-9]{2})\b")
# Structural tokens + the pack's own acronyms (`domain_acronyms`) — one
# derivation for planner/validator/writer (genericisation sweep; the three
# hand-copied UPI lists drifted: only the planner knew PSP1).
_ERROR_CODE_STOPWORDS: frozenset[str] = domain_vocab.error_code_stopwords()
_API_TOKEN_RE = re.compile(r"\b(?:Req|Resp)[A-Z][A-Za-z0-9]{2,}\b")


def _testcase_layout_check(sheet: SheetSpec, ws, plan: WorkbookPlan) -> list[Defect]:
    defects: list[Defect] = []
    layout = LAYOUT_REGISTRY[sheet.layout]
    headers = [ws.cell(layout.start_row, layout.start_col + idx).value for idx in range(len(layout.headers))]
    if headers != layout.headers:
        defects.append(Defect(
            severity="critical", sheet=ws.title, row=layout.start_row, type="header_mismatch",
            message=f"Header row mismatch. Expected {layout.headers}, got {headers}.",
            fix_hint="Re-render with the expected layout.",
        ))

    test_id_col = layout.start_col if sheet.layout in {"A1", "B1"} else layout.start_col + 1
    seen: dict[str, int] = {}
    suffixes_by_prefix: dict[str, list[int]] = defaultdict(list)

    for offset in range(len(sheet.test_cases)):
        row = layout.start_row + 1 + offset
        test_id = ws.cell(row, test_id_col).value
        if not test_id:
            defects.append(Defect(severity="critical", sheet=ws.title, row=row, type="empty_test_id",
                                  message="Missing TEST ID.", fix_hint="Populate a unique TEST ID."))
            continue
        test_id_str = str(test_id)
        if test_id_str in seen:
            defects.append(Defect(severity="critical", sheet=ws.title, row=row, test_id=test_id_str,
                                  type="duplicate_test_id", message=f"Duplicate TEST ID `{test_id_str}` (also row {seen[test_id_str]}).",
                                  fix_hint="Renumber the duplicate case."))
        else:
            seen[test_id_str] = row

        match = _TESTID_RE.match(test_id_str)
        if match:
            prefix = test_id_str.rsplit("_", 1)[0]
            suffixes_by_prefix[prefix].append(int(match.group(1)))

        for cell in ws[row]:
            if isinstance(cell.value, str) and "\n" in cell.value and not cell.alignment.wrap_text:
                defects.append(Defect(severity="warning", sheet=ws.title, row=row, test_id=test_id_str,
                                      type="wrap_text", message="Multiline cell is not wrapped.",
                                      fix_hint="Set wrap_text=True on this column."))

    prefix_counts = Counter(test_id.rsplit("_", 1)[0] for test_id in seen if "_" in test_id)
    if len(prefix_counts) > 1:
        common = prefix_counts.most_common(1)[0][0]
        for test_id, row in seen.items():
            if "_" in test_id and not test_id.startswith(common):
                defects.append(Defect(severity="warning", sheet=ws.title, row=row, test_id=test_id,
                                      type="prefix_mismatch",
                                      message=f"TEST ID prefix `{test_id.rsplit('_',1)[0]}` differs from sheet's dominant prefix `{common}`.",
                                      fix_hint=f"Renumber under prefix `{common}` for consistency."))

    for prefix, suffixes in suffixes_by_prefix.items():
        suffixes_sorted = sorted(set(suffixes))
        for prev, curr in zip(suffixes_sorted, suffixes_sorted[1:]):
            if curr - prev > 1:
                defects.append(Defect(severity="warning", sheet=ws.title, row=None, test_id=f"{prefix}_{curr}",
                                      type="test_id_gap",
                                      message=f"Gap in `{prefix}` test-id sequence between {prefix}_{prev} and {prefix}_{curr}.",
                                      fix_hint="Fill the gap or move case to Removed cases."))

    status_col_index = layout.start_col + next(
        i for i, h in enumerate(layout.headers) if h.strip().lower() == "status"
    )
    for offset in range(len(sheet.test_cases)):
        row = layout.start_row + 1 + offset
        status_value = ws.cell(row, status_col_index).value
        if status_value and str(status_value) not in _VALID_STATUSES:
            defects.append(Defect(severity="warning", sheet=ws.title, row=row,
                                  type="status_enum", message=f"Status `{status_value}` is not a recognised value.",
                                  fix_hint="Use Success/Failure/Deemed/etc per archetype casing."))

    pair_groups: dict[str, list[tuple[int, str, str]]] = defaultdict(list)
    detail_col_index = layout.start_col + 1 if sheet.layout in ("A1", "B1") else layout.start_col + 2
    desc_col_index = detail_col_index + 1
    if sheet.layout == "C2":
        desc_col_index += 1
    for offset, tc in enumerate(sheet.test_cases):
        if not tc.pair_id:
            continue
        row = layout.start_row + 1 + offset
        details = ws.cell(row, detail_col_index).value or ""
        desc = ws.cell(row, desc_col_index).value or ""
        pair_groups[tc.pair_id].append((row, details, desc))

    for pair_id, members in pair_groups.items():
        if len(members) < 2:
            continue
        first_details = members[0][1]
        first_desc = members[0][2]
        for row, details, desc in members[1:]:
            if details != first_details or desc != first_desc:
                defects.append(Defect(severity="critical", sheet=ws.title, row=row,
                                      type="pair_drift",
                                      message=f"Pair `{pair_id}` rows have different DETAILS or DESCRIPTION.",
                                      fix_hint="Copy DETAILS+DESCRIPTION from the partner row verbatim."))

    return defects


def _coverage_audit_check(plan: WorkbookPlan) -> list[Defect]:
    """Cross-check plan.coverage_audit vs the actual stub counts."""
    actual: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for sheet in plan.sheets:
        for tc in sheet.test_cases:
            for api in tc.apis:
                if not api.startswith("Req"):
                    continue
                actual[api][tc.coverage_tag] += 1

    defects: list[Defect] = []
    for api, audit in plan.coverage_audit.items():
        for tag, declared in audit.items():
            real = actual.get(api, {}).get(tag, 0)
            if real != declared:
                defects.append(Defect(severity="critical", sheet="(plan)", row=None,
                                      type="coverage_drift",
                                      message=f"coverage_audit declares {api}.{tag}={declared} but rendered count is {real}.",
                                      fix_hint="Recompute coverage_audit from rendered test cases."))
    return defects


def _home_link_check(wb, plan: WorkbookPlan) -> list[Defect]:
    if plan.archetype != "C":
        return []
    defects = []
    for sheet in plan.sheets:
        if sheet.layout not in {"C1", "C2"}:
            continue
        ws = wb[sheet.name[:31]] if sheet.name[:31] in wb.sheetnames else None
        if ws is None:
            continue
        cell = ws["B1"]
        if cell.value != "HOME" or not cell.hyperlink:
            defects.append(Defect(severity="warning", sheet=ws.title, row=1,
                                  type="home_link_broken",
                                  message="Archetype C sheet missing HOME hyperlink in B1.",
                                  fix_hint="Set B1 to 'HOME' with hyperlink #'Index'!A1."))
    return defects


def _merged_range_check(ws) -> list[Defect]:
    defects: list[Defect] = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row != merged_range.max_row:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                if merged_range.max_row - merged_range.min_row >= 1 and merged_range.min_row > 2:
                    defects.append(Defect(severity="warning", sheet=ws.title, row=row,
                                          type="merged_data_rows",
                                          message=f"Merged range {merged_range} crosses data rows.",
                                          fix_hint="Avoid merging cells that span test-case rows."))
                    break
    return defects


def _formula_required_check(wb, plan: WorkbookPlan) -> list[Defect]:
    defects: list[Defect] = []
    if plan.archetype != "B":
        return defects
    if "Version Log" not in wb.sheetnames:
        return defects
    ws = wb["Version Log"]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower().startswith("total"):
                next_cell = ws.cell(cell.row, cell.column + 1)
                if next_cell.value is not None and not (isinstance(next_cell.value, str) and next_cell.value.startswith("=")):
                    defects.append(Defect(severity="warning", sheet=ws.title, row=cell.row,
                                          type="formula_required",
                                          message="Version Log Total row should use =SUM(...) instead of a hardcoded value.",
                                          fix_hint="Replace with a SUM formula over the test-case-count column."))
    return defects


def _allowlists_from_engine_context(
    engine_context: dict | None, plan: WorkbookPlan,
) -> tuple[frozenset[str], frozenset[str]]:
    """Build (api_allowlist, code_allowlist) from the TSD interface_spec and
    any BRD/TSD error codes visible, plus what the plan actually declares.
    """
    apis: set[str] = set()
    codes: set[str] = {"00"}
    if engine_context:
        tsd = engine_context.get("tsd_sections") or {}
        apis.update(_API_TOKEN_RE.findall(tsd.get("interface_spec") or ""))
        for tok in _ERROR_CODE_RE.findall(tsd.get("error_handling") or ""):
            if tok not in _ERROR_CODE_STOPWORDS:
                codes.add(tok)
    for sheet in plan.sheets:
        for tc in sheet.test_cases:
            apis.update(tc.apis)
            if tc.response_code:
                codes.add(tc.response_code)
    return frozenset(apis), frozenset(codes)


def _step_lint_defects(
    plan: WorkbookPlan, engine_context: dict | None = None,
) -> list[Defect]:
    """Promote step-linter findings to Defects so they show up in the report.

    BRD/TSD-only: allowlist comes from the TSD interface_spec plus what the
    plan itself declares — no canonical UPI catalog.
    """
    extra_apis, extra_codes = _allowlists_from_engine_context(engine_context, plan)
    out: list[Defect] = []
    report = lint_plan(plan, extra_apis=extra_apis, extra_codes=extra_codes)
    severity_map = {
        "invalid_api_in_steps": "critical",
        "api_not_in_details": "critical",
        "missing_failure_code": "critical",
        "failure_code_mismatch": "critical",
        "invalid_error_code": "warning",
        "step_numbering": "warning",
        "success_terminus_missing": "warning",
        "entity_not_in_details": "warning",
        "cred_block_missing": "warning",
        "not_rendered": "critical",
    }
    sheet_for_test: dict[str, str] = {}
    for sheet in plan.sheets:
        for tc in sheet.test_cases:
            sheet_for_test[tc.test_id] = sheet.name
    for test_id, issues in report.issues_by_test.items():
        for issue in issues:
            severity = severity_map.get(issue.code, "warning")
            out.append(Defect(
                severity=severity,  # type: ignore[arg-type]
                sheet=sheet_for_test.get(test_id, "(unknown)"),
                row=None,
                test_id=test_id,
                type=issue.code,
                message=issue.message,
                fix_hint=issue.fix_hint,
            ))
    return out


def mechanical_check(
    plan: WorkbookPlan, path: Path, engine_context: dict | None = None,
) -> list[Defect]:
    """Run all mechanical structural + content checks plus the step linter."""
    defects: list[Defect] = []
    wb = load_workbook(path, data_only=False)

    for error in scan_formula_errors(path):
        defects.append(Defect(severity="critical", sheet=error.sheet, row=None, type="formula_error",
                              message=f"{error.cell}: {error.value}", fix_hint="Fix or remove broken formula."))

    for sheet in plan.sheets:
        if sheet.layout not in LAYOUT_REGISTRY:
            continue
        title = sheet.name[:31]
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        defects.extend(_testcase_layout_check(sheet, ws, plan))
        defects.extend(_merged_range_check(ws))

    defects.extend(_home_link_check(wb, plan))
    defects.extend(_coverage_audit_check(plan))
    defects.extend(_formula_required_check(wb, plan))
    defects.extend(_step_lint_defects(plan, engine_context))

    LOGGER.info("validator.mechanical", critical=sum(1 for d in defects if d.severity == "critical"),
                warning=sum(1 for d in defects if d.severity == "warning"))
    return defects


def _read_sheet_rows(path: Path, sheet_name: str) -> list[list]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name[:31] not in wb.sheetnames:
        return []
    ws = wb[sheet_name[:31]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


async def _semantic_check_one_sheet(sheet: SheetSpec, path: Path) -> list[Defect]:
    client = get_client("validator")
    rows = _read_sheet_rows(path, sheet.name)
    if not rows:
        return []
    system = [SystemBlock(text=load_prompt("validator.md"), cache=True)]
    user_msg = (
        f"Sheet name: {sheet.name}\nLayout: {sheet.layout}\n\n"
        f"Rows (header + body):\n{rows}\n\n"
        "Return a JSON object with key `defects` (list)."
    )
    try:
        response = await client.complete(
            system=system,
            messages=[Message(role="user", content=user_msg)],
            max_tokens=4000,
            response_format="json",
        )
        payload = parse_json_response(response.text)
        if isinstance(payload, list):
            payload = {"defects": payload}
        return [Defect.model_validate(d) for d in payload.get("defects", [])]
    except (ValidationError, ValueError) as exc:
        LOGGER.warning("validator.semantic_invalid_json", sheet=sheet.name, error=repr(exc)[:200])
        return []
    except Exception as exc:  # noqa: BLE001
        LOGGER.error("validator.semantic_error", sheet=sheet.name, error=repr(exc))
        return []


async def semantic_check(plan: WorkbookPlan, path: Path) -> list[Defect]:
    """Parallel LLM checks per sheet. Bounded by config.validator.max_concurrent_sheets."""
    runtime = load_runtime_config()
    target = [s for s in plan.sheets if s.layout in LAYOUT_REGISTRY]
    if not target:
        return []

    sem = asyncio.Semaphore(runtime.validator.max_concurrent_sheets)

    async def run(sheet: SheetSpec) -> list[Defect]:
        async with sem:
            return await _semantic_check_one_sheet(sheet, path)

    results = await asyncio.gather(*[run(s) for s in target])
    flat: list[Defect] = []
    for batch in results:
        flat.extend(batch)
    return flat


async def validate(
    plan: WorkbookPlan, path: Path, options: dict | None = None,
) -> ValidationReport:
    """Run mechanical and semantic validation in parallel.

    BRD/TSD-only: ``options["tsd_sections"]`` extends the step-lint allowlist.
    """
    mech, sem = await asyncio.gather(
        asyncio.to_thread(mechanical_check, plan, path, options),
        semantic_check(plan, path),
    )
    defects = mech + sem
    has_critical = any(d.severity == "critical" for d in defects)
    LOGGER.info("validator.done", critical=sum(1 for d in defects if d.severity == "critical"),
                warning=sum(1 for d in defects if d.severity == "warning"))
    return ValidationReport(status="fail" if has_critical else "pass", defects=defects)

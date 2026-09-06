# Copyright 2026 National Payments Corporation of India
# SPDX-License-Identifier: MIT

# INTEGRATION: FastAPI router with the engine-specific download endpoints.
#
# WHY these endpoints: Markdown and DOCX historically flow through the
# host's existing artifact-download path (change_requests.py:360+) which
# reads the ProductKitDocument table. The host's frontend wires those
# buttons. The bits that ARE engine-specific:
#   - the .xlsx — NPCI-format workbook, only the engine renders it.
#   - the .json — fully-typed WorkbookPlan dump for downstream tooling.
#   - user-upload override for the .xlsx (see /upload, /revert below).
# We add only those routes; everything else is shared.

from __future__ import annotations

import hashlib
import uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import FileResponse

from app.core.deps import CurrentUser, DbDep
from app.core.error_taxonomy import client_safe_detail
from app.excel_testcase_engine.adapters import jobs as jobs_adapter
from app.excel_testcase_engine.observability import get_logger

LOGGER = get_logger("excel_engine.api")

router = APIRouter(prefix="/changes", tags=["excel_testcase_engine"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
JSON_MIME = "application/json"

# 25 MB cap on user-uploaded xlsx — matches the video-upload cap noted in
# change_dispatch._doc_payload (line 115) and is well above any realistic
# certification workbook size (largest gold-example is ~350 KB).
_MAX_XLSX_UPLOAD_BYTES = 25 * 1024 * 1024


def _latest_succeeded_cert_job(db, change_id: str):
    """Pick the most-recent SUCCEEDED cert_test_cases job for this change.
    Shared between the .xlsx and .json download endpoints."""

    from app.models.agent_job import AgentJob, AgentJobStatus

    return (
        db.query(AgentJob)
        .filter(
            AgentJob.change_request_id == change_id,
            AgentJob.module == "product_kit",
            AgentJob.subtype == "cert_test_cases",
            AgentJob.status == AgentJobStatus.SUCCEEDED,
        )
        .order_by(AgentJob.completed_at.desc().nullslast(), AgentJob.updated_at.desc())
        .first()
    )


def _serve_companion(db, change_id: str, current_user, *, kind: str, mime: str, ext: str):
    """Common implementation for cert_test_cases companion downloads.

    `kind` is the key inside result_payload.files (xlsx / md / docx / json).
    """
    # A cert workbook can carry sensitive UPI flow detail — gate the download to
    # creator/admin/BRD-approver, the same rule upload/revert enforce (not just
    # any authenticated user).
    _access_check(db, change_id, current_user)

    job = _latest_succeeded_cert_job(db, change_id)
    if job is None:
        raise HTTPException(
            status_code=404,
            detail=(
                "No completed cert_test_cases workbook for this change request. "
                "Generate one first from the Product Kit panel."
            ),
        )

    result = job.result_payload or {}
    files = result.get("files") or {}
    legacy_keys = {
        "xlsx": "xlsx_path",
        "md": "md_path",
        "docx": "docx_path",
        "json": "json_path",
    }
    # Back-compat: older succeeded jobs stored companion paths at the top
    # level rather than under result_payload.files.
    file_path = files.get(kind) or result.get(legacy_keys.get(kind, ""))
    if not file_path:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Workbook record exists but is missing its {kind} path "
                f"(job={job.id[:12]}). Re-run generation to repair."
            ),
        )
    p = Path(file_path)
    if not p.exists():
        # WHY 410 (Gone) rather than 404: the record says we made the file
        # but it's not on disk — distinguishes "no workbook ever" from
        # "workbook was here, then rotated/cleaned".
        LOGGER.warning(f"{kind}.download_missing_file", job_id=job.id, path=file_path)
        raise HTTPException(
            status_code=410,
            detail=f"Workbook {kind} file is no longer available. Regenerate cert test cases to recreate it.",
        )

    version = result.get("version") or 1
    filename = f"cert_testcases_{change_id[:8]}_v{version}.{ext}"
    return FileResponse(p, filename=filename, media_type=mime)


@router.get("/{change_id}/product-kit/cert_test_cases/xlsx")
async def download_cert_testcase_xlsx(
    change_id: str,
    current_user: CurrentUser,
    db: DbDep,
):
    """Download the most-recently-completed cert_test_cases NPCI workbook.
    Access-gated to creator/admin/BRD-approver (see _serve_companion)."""
    return _serve_companion(db, change_id, current_user, kind="xlsx", mime=XLSX_MIME, ext="xlsx")


@router.get("/{change_id}/product-kit/cert_test_cases/json")
async def download_cert_testcase_json(
    change_id: str,
    current_user: CurrentUser,
    db: DbDep,
):
    """Download the most-recently-completed cert_test_cases WorkbookPlan as
    JSON. Useful for downstream tooling that needs structured access to the
    test cases without parsing the .xlsx (test-management imports, diffs,
    inspection, automated downstream pipelines).

    Fast path: return the pre-computed json companion (`files.json`) as-is.
    Fallback: when the latest cert job has an xlsx but NO json companion
    (legacy engine rows / uploads that predate the JSON-parsing feature /
    a companion that was cleaned off disk), parse the xlsx on-demand into
    a WorkbookPlan-lite dict so the table view keeps rendering rather than
    silently 404-ing. Self-heals the stale-state scenario without needing a
    DB backfill.
    """
    from fastapi.responses import JSONResponse, Response
    import json as _json

    # Same access rule as the xlsx download / upload / revert — not just any
    # authenticated user (a cert workbook can carry sensitive UPI flow detail).
    _access_check(db, change_id, current_user)

    job = _latest_succeeded_cert_job(db, change_id)
    if job is None:
        raise HTTPException(
            status_code=404,
            detail=(
                "No completed cert_test_cases workbook for this change request. "
                "Generate one first from the Product Kit panel."
            ),
        )
    result = job.result_payload or {}
    files = result.get("files") or {}
    json_path = files.get("json") or result.get("json_path")
    xlsx_path = files.get("xlsx") or result.get("xlsx_path")

    # Cache-Control: no-store on ALL responses from this endpoint.
    # The upload/revert flow can rewrite the "latest cert job" between two
    # otherwise-identical URL requests, so the browser MUST refetch every
    # time — a cached response is nearly guaranteed to be stale.
    no_store_headers = {"Cache-Control": "no-store, no-cache, must-revalidate"}

    # Fast path: pre-computed json companion still on disk → serve it.
    if json_path and Path(json_path).exists():
        version = result.get("version") or 1
        filename = f"cert_testcases_{change_id[:8]}_v{version}.json"
        return FileResponse(
            Path(json_path), filename=filename, media_type=JSON_MIME,
            headers=no_store_headers,
        )

    # Fallback: parse the xlsx on the fly. Requires xlsx to exist.
    if xlsx_path and Path(xlsx_path).exists():
        try:
            data = Path(xlsx_path).read_bytes()
            plan = _xlsx_to_workbook_plan_json(
                data, filename=Path(xlsx_path).name,
            )
            LOGGER.info(
                "cert_test_cases.json_fallback_parsed",
                job_id=job.id, change_id=change_id,
                test_cases=len(plan.get("test_cases") or []),
            )
            body = _json.dumps(plan, ensure_ascii=False).encode("utf-8")
            return Response(
                content=body, media_type=JSON_MIME,
                headers=no_store_headers,
            )
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning(
                "cert_test_cases.json_fallback_failed",
                job_id=job.id, error=repr(exc)[:200],
            )
            # Fall through to the 404 below with a diagnostic detail.

    raise HTTPException(
        status_code=404,
        detail=(
            f"Workbook record exists but is missing its json companion "
            f"(job={job.id[:12]}). Re-run generation or re-upload to repair."
        ),
    )


# ── User-upload override ───────────────────────────────────────────────────
#
# Design: rather than mutate an existing agent job or add a new column, we
# insert a NEW `AgentJob` row with `module=product_kit`, `subtype=cert_test_cases`,
# `status=SUCCEEDED`, whose `result_payload.files.xlsx` points at the saved
# upload. Because `_latest_succeeded_cert_job` orders by
# `completed_at DESC, updated_at DESC`, this row becomes the source of truth
# for BOTH the download endpoint above AND `change_dispatch.build_kit_envelope`
# — no changes needed to either. The prior job(s) remain in the DB as an
# audit trail; `/revert` inserts another superseding job that re-points at
# the last non-user-upload xlsx.


def _access_check(db, change_id: str, current_user):
    """Match the download-side access rule: creator, admin, or an approver on
    this change's BRD. Mirrors change_requests.download_artifact:746 but
    scoped to cert-testcase upload/revert only (no product-kit-review flag —
    review teams can VIEW but not edit the shipping xlsx)."""
    from sqlalchemy import func, select
    from app.models.change_request import ChangeRequest

    change = db.get(ChangeRequest, change_id)
    if not change:
        raise HTTPException(status_code=404, detail="Change request not found")

    # Local import to match _can_read_all_changes / _is_creator_or_admin
    # pattern in change_requests.py — avoids circular imports at module load.
    from app.models.user import UserRole
    if current_user.role == UserRole.ADMIN or change.created_by == current_user.id:
        return change

    # Approver on the change's BRD — same query change_requests.py runs.
    from app.models.approval import Approval
    from app.models.brd import BRD
    is_approver = (db.scalar(
        select(func.count(Approval.id)).where(
            Approval.approver_id == current_user.id,
            Approval.artifact_id.in_(
                select(BRD.id).where(BRD.change_request_id == change_id)
            ),
        )
    ) or 0) > 0
    if not is_approver:
        raise HTTPException(status_code=403, detail="Access denied")
    return change


def _validate_xlsx_bytes(data: bytes) -> list[str]:
    """Open the uploaded bytes as an xlsx. Return a list of shape-warning
    strings (empty when everything looks clean). Raises HTTPException(415)
    when the bytes aren't a parseable xlsx at all — that's the only hard
    reject; downstream shape drift is a warning, not a block, so the user
    can push a genuinely fixed pack even when column headers moved."""
    try:
        import openpyxl  # noqa: WPS433 — deferred import; heavy dep
        from openpyxl.utils.exceptions import InvalidFileException
    except Exception as exc:  # noqa: BLE001
        # SCR #6: an ImportError message can disclose interpreter/filesystem
        # layout. The cause is preserved for operators via `from exc`.
        raise HTTPException(
            status_code=500,
            detail="Server missing openpyxl for upload validation",
        ) from exc

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except (InvalidFileException, Exception) as exc:  # noqa: BLE001
        # SCR #6: openpyxl raises a bare KeyError for a zip that is not a
        # workbook, and its text quotes the UPLOADER'S OWN archive member name
        # back to them — attacker-controlled content reflected into a response.
        # InvalidFileException carries a message we can show; anything else is
        # reduced to a category label by client_safe_detail().
        LOGGER.warning("xlsx upload rejected: %s: %s", type(exc).__name__, exc)
        raise HTTPException(
            status_code=415,
            detail=f"Uploaded file is not a valid .xlsx workbook: {client_safe_detail(exc)}",
        ) from exc

    warnings: list[str] = []
    sheet_names = wb.sheetnames
    if not sheet_names:
        warnings.append("Workbook has no sheets.")
    # Shape drift: presence of an "Index" or role sheet is a strong signal
    # this is an NPCI-format pack. Absence isn't fatal — but flag it so the
    # reviewer knows the pack may not slot into downstream tooling.
    from app.excel_testcase_engine import domain_vocab
    known_sheets = ("Index", "Summary", "Sheet1", "Version Log",
                    *domain_vocab.role_sheet_names(), *domain_vocab.party_labels())
    pack_ish = any(s in sheet_names for s in known_sheets)
    if not pack_ish:
        warnings.append(
            "Sheet names do not match any known certification-pack layout "
            f"(saw {sheet_names[:6]}); downstream partner tooling may not parse this."
        )
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    return warnings


def _xlsx_to_markdown(data: bytes, *, max_cell_chars: int = 4000) -> str:
    """Render an uploaded xlsx as markdown so the Certification Test Cases
    panel can display it via ReactMarkdown.

    Not a WorkbookPlan round-trip — that would require reversing the entire
    Writer + renderer chain. Instead we produce a faithful per-sheet /
    per-row dump: every sheet becomes a `## <name>` section, every non-empty
    row becomes a `### <sheet>_<row>` subsection with one bullet per column.

    Multi-line cell values (DETAILS blocks, TEST STEPS) are wrapped in a
    fenced code block so their `\\n` breaks render intact — plain markdown
    tables can't carry newlines.
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = [
        "> _This workbook was uploaded by a user. It replaces the previously "
        "generated pack; the download button and Product Kit shipping now "
        "use this version._\n",
    ]

    for sheet in wb.worksheets:
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        # Drop trailing blank rows (openpyxl often reports many).
        while rows and not any(_cell_is_populated(c) for c in rows[-1]):
            rows.pop()
        if not rows:
            continue

        parts.append(f"## {sheet.title}\n")

        # Assume the first row is a header (real NPCI packs always have one).
        headers = [
            (str(h).strip() if _cell_is_populated(h) else f"Col{i+1}")
            for i, h in enumerate(rows[0])
        ]

        for row_idx, row in enumerate(rows[1:], start=2):
            # Skip blank rows in the middle (separator rows).
            if not any(_cell_is_populated(c) for c in row):
                continue

            # Pick a stable row title: prefer the leftmost populated cell
            # (typically the TEST ID column).
            row_label = ""
            for cell in row:
                if _cell_is_populated(cell):
                    row_label = str(cell).strip().splitlines()[0][:80]
                    break
            parts.append(f"### {sheet.title} · row {row_idx} — {row_label}\n")

            for header, cell in zip(headers, row):
                if not _cell_is_populated(cell):
                    continue
                value = str(cell)
                if len(value) > max_cell_chars:
                    value = value[:max_cell_chars] + "\n… (truncated)"
                # Multi-line cells go in a fenced code block; short cells
                # go inline. Preserves DETAILS / TEST STEPS formatting.
                if "\n" in value:
                    parts.append(f"**{header}**\n\n```\n{value}\n```\n")
                else:
                    parts.append(f"- **{header}**: {value}\n")
            parts.append("")  # blank line between rows

    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    return "\n".join(parts).rstrip() + "\n"


def _cell_is_populated(value) -> bool:
    """Openpyxl returns None for empty cells and sometimes empty strings
    with only whitespace; both count as blank."""
    if value is None:
        return False
    return bool(str(value).strip())


# ── xlsx → WorkbookPlan-lite JSON (drives the CertTestCasesTable view) ────


# Recognised column headers, case-insensitive. Multiple synonyms per field
# because real NPCI packs vary in casing/naming across archetypes A/B/C and
# across teams. Missing → the corresponding field is left empty.
#
# Deliberately NOT synonyms of `test_id`:
#   "s.no", "sno" — those are SERIAL numbers, not test-case IDs. Some packs
#   have both a S.NO column and a TEST ID column; treating S.NO as test_id
#   would clobber the real test_id.
_COL_SYNONYMS = {
    "test_id":          ("test id", "tc id", "test case id", "tcid", "testcase id"),
    "expected_status":  ("status", "expected status", "test status", "result", "expected result"),
    "response_code":    ("response code", "expected response code", "resp code", "resp",
                         "expected resp", "error code", "code"),
    "api_type":         ("type", "test type", "flow type"),
    "coverage_tag":     ("coverage", "coverage tag", "tag"),
    "scenario_summary": ("scenario", "scenario summary", "test scenario", "test case scenario",
                         "test case", "scenario name"),
    "details_block":    ("details", "test details", "test case details", "detail"),
    "description_block":("description", "test description", "test case description", "desc"),
    "steps_block":      ("test steps", "steps", "test procedure", "procedure",
                         "steps to execute", "execution steps"),
    "txn_initiated_by": ("txn initiated by", "initiated by", "transaction initiated by"),
    # The "psp as" spellings are UPI's historical column names; the active
    # pack's own `role_as_label` (e.g. "Library as") is appended at match
    # time in `_alias_map()` below.
    "psp_as":           ("psp as", "psp role", "psp acting as"),
    "apis":             ("apis", "api involved", "api list", "api"),
    "priority":         ("priority", "prio"),
    "traces":           ("traces", "traceability", "brd trace", "brd ref"),
    "covers_field":     ("feature field", "field", "covers field"),
    "covers_business_rule": ("business rule", "covered rule", "covers business rule"),
    "scope":            ("scope",),
}


def _canonical_field(header: str) -> str | None:
    """Map a raw header cell to a canonical field name, or None."""
    h = (header or "").strip().lower()
    if not h:
        return None
    # The active pack's own role-as column label ("PSP as" / "Library as")
    # maps to `psp_as` alongside UPI's historical spellings.
    from app.core.domain.registry import prompt_block
    role_as = prompt_block("role_as_label", "").strip().lower()
    if role_as and h == role_as:
        return "psp_as"
    for field, syns in _COL_SYNONYMS.items():
        if h in syns:
            return field
    return None


_STATUS_NORMALISE = {
    "success": "Success",
    "pass":    "Success",
    "passed":  "Success",
    "failure": "Failure",
    "fail":    "Failure",
    "failed":  "Failure",
    "deemed":  "Deemed",
    "partial": "Partial",
    "pending": "Failure",  # rare; treat as failure so the pill still colours
}


def _normalise_status(value: str) -> str:
    v = (value or "").strip().lower()
    return _STATUS_NORMALISE.get(v, value or "")


# Regex to pull API names out of a DETAILS block ("API Involved : ReqTransfer,\n
# RespTransfer\nType : Pay\n…"). Deliberately conservative — anything not shaped
# like a bare Req*/Resp* token is skipped.
import re as _re
_API_LINE_RE = _re.compile(r"api\s+involved\s*:\s*([^\n\r]+(?:\n\s+[^\n\r]+)*)", _re.IGNORECASE)
_API_TOKEN_RE = _re.compile(r"\b(Req\w+|Resp\w+)\b")


def _apis_from_details(details_text: str) -> list[str]:
    if not details_text:
        return []
    m = _API_LINE_RE.search(details_text)
    if not m:
        return []
    return _API_TOKEN_RE.findall(m.group(1))


def _apis_from_cell(value: str) -> list[str]:
    """Split a comma / newline delimited APIs cell into a list."""
    if not value:
        return []
    parts: list[str] = []
    for chunk in _re.split(r"[,\n\r/]+", str(value)):
        token = chunk.strip()
        if token:
            parts.append(token)
    return parts


def _detect_header_row(rows: list[list], *, scan_depth: int = 6) -> tuple[int, dict[int, str]]:
    """Find the row that best looks like a column-header row.

    Real NPCI annexure workbooks often have decorative rows at the top —
    a "HOME" navigation link, a banner, a merged title. The real column
    headers (TEST ID / DETAILS / STATUS / TEST STEPS) sit at row 1 or 2,
    not row 0. We scan the first `scan_depth` rows and pick the one whose
    cells match the most `_COL_SYNONYMS` entries — that's the header.

    Returns (header_row_index, col_map). Both empty when no header found.
    """
    best_row_idx = -1
    best_col_map: dict[int, str] = {}
    for row_idx, row in enumerate(rows[:scan_depth]):
        col_map: dict[int, str] = {}
        for cell_idx, cell in enumerate(row):
            if not _cell_is_populated(cell):
                continue
            field = _canonical_field(str(cell))
            if field:
                col_map[cell_idx] = field
        # More recognised headers wins. Ties: earlier row wins (typical
        # case where the same header shape repeats due to merged banners).
        if len(col_map) > len(best_col_map):
            best_col_map = col_map
            best_row_idx = row_idx
    return best_row_idx, best_col_map


# Sheet-name substrings that indicate a metadata sheet, not a test-case
# sheet. Match is case-insensitive substring, so "SUMMARY SHEET1" catches
# via "summary" and "MODES OF CERTIFICATION1" via "modes of certification".
# Deliberately conservative — false negatives (letting a case sheet through)
# are cheaper than false positives (dropping a real role sheet).
_METADATA_SHEET_HINTS = (
    "index",
    "summary",
    "subset",
    "modes of certification",
    "version history",
    "scope",
    "validation_report",
    "validation report",
    "changelog",
    "cover",
    "revision history",
    "toc",
    "table of contents",
    "notes",
    "legend",
)


def _is_metadata_sheet(name: str) -> bool:
    n = (name or "").strip().lower()
    if not n:
        return True
    return any(hint in n for hint in _METADATA_SHEET_HINTS)


def _xlsx_to_workbook_plan_json(data: bytes, *, filename: str = "uploaded.xlsx") -> dict:
    """Parse an uploaded xlsx into a WorkbookPlan-shaped dict that the
    CertTestCasesTable view can render.

    Not a strict WorkbookPlan validation — we produce a *lite* shape:
    every sheet becomes a `sheets[]` entry with `test_cases[]`, and the
    top-level `test_cases` flat list carries every case across sheets so
    the existing table (which iterates `plan.test_cases`) keeps working
    without a schema change.

    Unknown columns are captured in the DETAILS block so nothing the
    reviewer added is lost.
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets_json: list[dict] = []
    flat_cases: list[dict] = []
    tc_counter = 0

    for sheet in wb.worksheets:
        # Skip known metadata sheets early — they occasionally contain
        # column names that fool _detect_header_row (e.g. a summary that
        # tables "STATUS" and "SCOPE" columns for merged banners).
        if _is_metadata_sheet(sheet.title):
            sheets_json.append({"name": sheet.title, "layout": "metadata", "test_cases": []})
            continue
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        while rows and not any(_cell_is_populated(c) for c in rows[-1]):
            rows.pop()
        if len(rows) < 2:
            continue  # header-only or empty sheet

        header_idx, col_map = _detect_header_row(rows)
        if header_idx < 0 or not col_map:
            sheets_json.append({"name": sheet.title, "layout": "", "test_cases": []})
            continue

        headers_raw = [
            str(h).strip() if _cell_is_populated(h) else ""
            for h in rows[header_idx]
        ]

        # Sheets whose header row has no test-case-ish columns are metadata
        # (Index / Summary / Modes of Certification in Archetype C) — skip
        # so their rows don't pollute the test-case table.
        recognised = set(col_map.values())
        if not (recognised & {"test_id", "scenario_summary", "details_block", "steps_block"}):
            sheets_json.append({"name": sheet.title, "layout": "", "test_cases": []})
            continue

        sheet_cases: list[dict] = []
        for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any(_cell_is_populated(c) for c in row):
                continue

            # Assemble the case dict field-by-field.
            case: dict = {
                "test_id": "", "expected_status": "", "response_code": "",
                "api_type": "", "coverage_tag": "", "scenario_summary": "",
                "txn_initiated_by": "", "psp_as": "",
                "apis": [], "entities": [],
                "approval_type": "", "payer_handle": "", "payee_handle": "",
                "scope": "", "priority": None,
                "covers_field": None, "covers_business_rule": None,
                "rendered": {"details_block": "", "description_block": "", "steps_block": ""},
            }
            extras: list[str] = []  # unknown-column preservation for DETAILS

            for idx, value in enumerate(row):
                if not _cell_is_populated(value):
                    continue
                str_val = str(value).strip("\n").rstrip()
                field = col_map.get(idx)
                if field is None:
                    extras.append(f"{headers_raw[idx]}: {str_val}")
                    continue
                if field == "test_id":
                    case["test_id"] = str(value).strip().splitlines()[0]
                elif field == "expected_status":
                    case["expected_status"] = _normalise_status(str_val)
                elif field == "apis":
                    case["apis"] = _apis_from_cell(str_val)
                elif field == "details_block":
                    case["rendered"]["details_block"] = str_val
                elif field == "description_block":
                    case["rendered"]["description_block"] = str_val
                elif field == "steps_block":
                    case["rendered"]["steps_block"] = str_val
                elif field == "traces":
                    case["traceability"] = {"brd_refs": [s.strip() for s in str_val.split(",") if s.strip()]}
                elif field == "priority":
                    p = str_val.upper()
                    if p in ("P0", "P1", "P2"):
                        case["priority"] = p
                else:
                    case[field] = str_val

            # Backfill test_id / scenario when the column was missing so
            # the table still has something clickable.
            if not case["test_id"]:
                tc_counter += 1
                case["test_id"] = f"UPL_{tc_counter:03d}"
            if not case["apis"]:
                case["apis"] = _apis_from_details(case["rendered"]["details_block"])
            if not case["scenario_summary"] and case["rendered"]["description_block"]:
                # First non-empty line of description as a fallback summary.
                for line in case["rendered"]["description_block"].splitlines():
                    line = line.strip()
                    if line:
                        case["scenario_summary"] = line[:200]
                        break
            if not case["expected_status"]:
                case["expected_status"] = "Success"  # neutral default so the pill renders

            if extras:
                # Append the unknown columns to DETAILS so the reviewer's
                # additions aren't dropped — visible in row-expand view.
                d = case["rendered"]["details_block"]
                case["rendered"]["details_block"] = (
                    (d + "\n\n" if d else "")
                    + "\n".join(extras)
                )

            sheet_cases.append(case)
            flat_cases.append(case)

        sheets_json.append({
            "name": sheet.title, "layout": "", "test_cases": sheet_cases,
        })

    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass

    return {
        "filename":            filename,
        "archetype":           "A",
        "feature_name":        "",
        "sheets":              sheets_json,
        "test_cases":          flat_cases,   # flat list for the table view
        "global_conventions":  {},
        "coverage_audit":      {},
        "flow_definitions":    [],
        # Marker so the frontend can render "user-uploaded pack" hints when
        # it wants to; harmless when ignored.
        "source":              "user_upload",
    }


def _uploads_dir(change_id: str) -> Path:
    """Directory where user-uploaded overrides land. Under the engine's
    outputs root so the deployment volume already covers it (see
    injector.register_excel_testcase_engine)."""
    from app.excel_testcase_engine.orchestrator import graph as _graph
    root = _graph._OUTPUTS_DIR / "uploads" / change_id
    root.mkdir(parents=True, exist_ok=True)
    return root


def _bump_version(latest_result: dict | None) -> int:
    """Version counter carried on result_payload. Starts at 2 (the initial
    generated pack is implicitly v1). Monotonic across upload+revert cycles."""
    prev = 0
    try:
        prev = int((latest_result or {}).get("version") or 0)
    except (TypeError, ValueError):
        prev = 0
    return max(2, prev + 1)


def _insert_superseding_cert_job(
    db,
    *,
    change_id: str,
    user_id: str | None,
    prev_job_id: str | None,
    result_payload: dict,
) -> str:
    """Insert a new SUCCEEDED cert_test_cases AgentJob that becomes the
    latest-wins row for `_latest_succeeded_cert_job`. Used by both /upload
    and /revert so the mechanism is symmetric."""
    from app.models.agent_job import AgentJob, AgentJobStatus

    now = datetime.now(timezone.utc)
    new_id = uuid.uuid4().hex
    job = AgentJob(
        id=new_id,
        change_request_id=change_id,
        module="product_kit",
        subtype="cert_test_cases",
        status=AgentJobStatus.SUCCEEDED,
        started_at=now,
        completed_at=now,
        updated_at=now,
        started_by_user_id=user_id,
        current_stage="Uploaded" if result_payload.get("source") == "user_upload" else "Reverted",
        progress_pct=100,
        result_payload=result_payload,
        metadata_={"supersedes_job": prev_job_id} if prev_job_id else {},
    )
    db.add(job)
    db.commit()
    LOGGER.info(
        "cert_test_cases.superseding_job",
        change_id=change_id, new_job=new_id, prev_job=prev_job_id,
        source=result_payload.get("source"),
        version=result_payload.get("version"),
    )
    return new_id


@router.post("/{change_id}/product-kit/cert_test_cases/xlsx/upload")
async def upload_cert_testcase_xlsx(
    change_id: str,
    current_user: CurrentUser,
    db: DbDep,
    file: UploadFile = File(...),
):
    """Upload a user-edited cert-testcase workbook. Becomes the new source of
    truth for downloads AND for Product Kit shipping — the prior generated
    pack is preserved as an AgentJob row for audit but is no longer the
    latest. Use /revert to fall back to the last generated pack.

    Response: {job_id, version, xlsx_size_bytes, uploaded_at,
               supersedes_job, warnings, sha256}.
    Status codes: 201 (created), 403 (access denied), 404 (change not found
    or no prior cert job), 413 (too large), 415 (not an xlsx).
    """
    _access_check(db, change_id, current_user)

    # Read + size-cap. UploadFile is a SpooledTemporaryFile — read once.
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(data) > _MAX_XLSX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"Upload exceeds cap: {len(data)} bytes > "
                f"{_MAX_XLSX_UPLOAD_BYTES} bytes"
            ),
        )

    warnings = _validate_xlsx_bytes(data)

    # Persist under the engine's outputs dir with a deterministic filename
    # (timestamp + sha1 prefix). Keeps prior uploads discoverable on disk.
    sha256 = hashlib.sha256(data).hexdigest()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    fname = f"{ts}_{sha256[:12]}.xlsx"
    target = _uploads_dir(change_id) / fname
    target.write_bytes(data)

    # Also derive a WorkbookPlan-lite JSON so the /xlsx/json companion
    # endpoint returns real data (drives the CertTestCasesTable). Without
    # this, the panel would fall back to markdown view after upload.
    json_target = target.with_suffix(".json")
    plan_json = _xlsx_to_workbook_plan_json(data, filename=file.filename or fname)
    import json as _json
    json_target.write_text(
        _json.dumps(plan_json, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    # Find the prior latest cert job (if any) — its version + id go into
    # the audit fields on the new row.
    prev = _latest_succeeded_cert_job(db, change_id)
    prev_id = prev.id if prev is not None else None
    prev_result = prev.result_payload if prev is not None else None
    new_version = _bump_version(prev_result)
    now = datetime.now(timezone.utc)

    result_payload = {
        "files": {"xlsx": str(target), "json": str(json_target)},
        "source": "user_upload",
        "uploaded_by": getattr(current_user, "id", None),
        "uploaded_at": now.isoformat(),
        "supersedes_job": prev_id,
        "version": new_version,
        "xlsx_size_bytes": len(data),
        "xlsx_sha256": sha256,
        "warnings": warnings,
        # Keep original filename for reviewer context — the served download
        # uses the deterministic filename above (cert_testcases_{id}_v{v}.xlsx).
        "original_filename": file.filename or "",
    }
    new_job_id = _insert_superseding_cert_job(
        db,
        change_id=change_id,
        user_id=getattr(current_user, "id", None),
        prev_job_id=prev_id,
        result_payload=result_payload,
    )

    # Also refresh the ProductKitDocument row that the "Certification Test
    # Cases" panel renders. Without this, the panel keeps showing the OLD
    # markdown even though the .xlsx download serves the uploaded file.
    # Uses source=UPLOADED — the frontend already treats that as "don't
    # try to fetch the WorkbookPlan JSON, just show the markdown" (see
    # ProductKit.jsx isDocUploaded gating on the cert_test_cases JSON query).
    try:
        _upsert_cert_kit_doc_from_xlsx(
            db,
            change_id=change_id,
            data=data,
            user_id=getattr(current_user, "id", None),
            source_value="uploaded",
        )
    except Exception as exc:  # noqa: BLE001
        # Defensive: the AgentJob + xlsx are already persisted, so failure
        # here means the download still works but the panel may show stale
        # markdown. Log-and-continue rather than roll back the successful
        # upload — user can retry.
        LOGGER.warning("cert_test_cases.pk_doc_upsert_failed", error=repr(exc)[:200])

    return {
        "job_id": new_job_id,
        "version": new_version,
        "xlsx_size_bytes": len(data),
        "uploaded_at": now.isoformat(),
        "supersedes_job": prev_id,
        "warnings": warnings,
        "sha256": sha256,
    }


def _upsert_cert_kit_doc_from_xlsx(
    db,
    *,
    change_id: str,
    data: bytes,
    user_id: str | None,
    source_value: str,
) -> str:
    """Insert a new ProductKitDocument row (source=UPLOADED) whose content
    is the uploaded xlsx rendered to markdown. Returns the row id.

    New row (not update-in-place) so history is preserved — matches the
    engine's own write pattern in api/agents.py:4056.
    """
    from datetime import datetime, timezone

    from app.models.document_source import DocumentSource
    from app.models.change_request import ChangeRequest
    from app.models.product_kit import ProductKitDocType, ProductKitDocument

    cr = db.get(ChangeRequest, change_id)
    nv = getattr(cr, "negotiation_version", 1) or 1

    # Version = max existing + 1 for (change, doc_type=cert_test_cases).
    latest = (
        db.query(ProductKitDocument)
        .filter(
            ProductKitDocument.change_request_id == change_id,
            ProductKitDocument.doc_type == ProductKitDocType.CERT_TEST_CASES,
        )
        .order_by(ProductKitDocument.version.desc())
        .first()
    )
    new_version = (latest.version if latest else 0) + 1

    markdown = _xlsx_to_markdown(data)
    now = datetime.now(timezone.utc)

    row = ProductKitDocument(
        change_request_id=change_id,
        doc_type=ProductKitDocType.CERT_TEST_CASES,
        content=markdown,
        version=new_version,
        negotiation_version=nv,
        source=DocumentSource(source_value) if source_value else DocumentSource.UPLOADED,
        uploaded_by=user_id,
        uploaded_at=now,
    )
    db.add(row)
    db.commit()
    LOGGER.info(
        "cert_test_cases.pk_doc_upserted",
        change_id=change_id, version=new_version,
        source=source_value, content_bytes=len(markdown),
    )
    return row.id


@router.post("/{change_id}/product-kit/cert_test_cases/xlsx/revert")
async def revert_cert_testcase_xlsx(
    change_id: str,
    current_user: CurrentUser,
    db: DbDep,
):
    """Revert to the last engine-generated cert-testcase workbook.

    Walks the cert-testcase job history and picks the most recent job whose
    result_payload.source is NOT 'user_upload' (i.e. the last engine-generated
    or already-reverted row). Inserts a new SUCCEEDED job that RE-POINTS at
    that xlsx path — same latest-wins mechanism as /upload. Returns 409 when
    there is no non-upload predecessor to revert to (nothing to fall back on).
    """
    _access_check(db, change_id, current_user)

    from app.models.agent_job import AgentJob, AgentJobStatus

    # Walk history newest → oldest, skipping the ones that are themselves
    # user uploads. First non-upload wins.
    history = (
        db.query(AgentJob)
        .filter(
            AgentJob.change_request_id == change_id,
            AgentJob.module == "product_kit",
            AgentJob.subtype == "cert_test_cases",
            AgentJob.status == AgentJobStatus.SUCCEEDED,
        )
        .order_by(AgentJob.completed_at.desc().nullslast(), AgentJob.updated_at.desc())
        .all()
    )
    target = None
    for job in history:
        rp = job.result_payload or {}
        if (rp.get("source") or "") != "user_upload":
            target = job
            break
    if target is None:
        raise HTTPException(
            status_code=409,
            detail=(
                "Nothing to revert to — no engine-generated cert_test_cases "
                "pack exists prior to the current upload. Regenerate from the "
                "Product Kit panel."
            ),
        )

    target_rp = target.result_payload or {}
    target_files = (target_rp.get("files") or {})
    target_xlsx = target_files.get("xlsx") or target_rp.get("xlsx_path")
    if not target_xlsx or not Path(target_xlsx).exists():
        raise HTTPException(
            status_code=410,
            detail=(
                f"Revert target xlsx is no longer on disk (job={target.id[:12]}). "
                "Regenerate the cert_test_cases pack."
            ),
        )

    prev = _latest_succeeded_cert_job(db, change_id)
    prev_id = prev.id if prev is not None else None
    prev_result = prev.result_payload if prev is not None else None
    new_version = _bump_version(prev_result)
    now = datetime.now(timezone.utc)

    result_payload = {
        # Copy the target's files dict so any md/docx/json companions stay
        # aligned — the revert restores the whole family, not just xlsx.
        "files": dict(target_files or {"xlsx": target_xlsx}),
        "source": "user_revert",
        "reverted_by": getattr(current_user, "id", None),
        "reverted_at": now.isoformat(),
        "supersedes_job": prev_id,
        "reverted_to_job": target.id,
        "version": new_version,
    }
    new_job_id = _insert_superseding_cert_job(
        db,
        change_id=change_id,
        user_id=getattr(current_user, "id", None),
        prev_job_id=prev_id,
        result_payload=result_payload,
    )

    # Restore the panel's markdown to the last engine-generated content so
    # revisiting Certification Test Cases shows the original pack, not the
    # last user upload. Mirrors _upsert_cert_kit_doc_from_xlsx but clones
    # from the ProductKitDocument history rather than re-parsing the xlsx
    # (parsing our own generic dump-format on top of an already-rendered
    # engine pack would look worse than the engine's own markdown).
    try:
        _clone_last_generated_kit_doc(
            db,
            change_id=change_id,
            user_id=getattr(current_user, "id", None),
        )
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("cert_test_cases.pk_doc_revert_failed", error=repr(exc)[:200])

    return {
        "job_id": new_job_id,
        "version": new_version,
        "reverted_to_job": target.id,
        "reverted_at": now.isoformat(),
        "supersedes_job": prev_id,
    }


def _clone_last_generated_kit_doc(
    db,
    *,
    change_id: str,
    user_id: str | None,
) -> str | None:
    """Clone the most-recent GENERATED-source ProductKitDocument into a new
    row with bumped version. Returns the new row id, or None when there is
    no generated predecessor to clone (nothing to revert to)."""
    from datetime import datetime, timezone

    from app.models.document_source import DocumentSource
    from app.models.change_request import ChangeRequest
    from app.models.product_kit import ProductKitDocType, ProductKitDocument

    # Walk newest → oldest, skip UPLOADED rows, first GENERATED wins.
    history = (
        db.query(ProductKitDocument)
        .filter(
            ProductKitDocument.change_request_id == change_id,
            ProductKitDocument.doc_type == ProductKitDocType.CERT_TEST_CASES,
        )
        .order_by(ProductKitDocument.version.desc())
        .all()
    )
    target = None
    for r in history:
        if r.source == DocumentSource.GENERATED:
            target = r
            break
    if target is None:
        return None

    new_version = (history[0].version if history else 0) + 1
    cr = db.get(ChangeRequest, change_id)
    nv = getattr(cr, "negotiation_version", 1) or 1
    now = datetime.now(timezone.utc)

    row = ProductKitDocument(
        change_request_id=change_id,
        doc_type=ProductKitDocType.CERT_TEST_CASES,
        content=target.content,   # clone the engine's original markdown verbatim
        version=new_version,
        negotiation_version=nv,
        source=DocumentSource.GENERATED,  # marked as generated so the JSON view re-enables
        uploaded_by=user_id,
        uploaded_at=now,
    )
    db.add(row)
    db.commit()
    LOGGER.info(
        "cert_test_cases.pk_doc_reverted",
        change_id=change_id, new_version=new_version,
        cloned_from_version=target.version,
    )
    return row.id


@router.get("/{change_id}/product-kit/cert_test_cases/xlsx/status")
async def cert_testcase_xlsx_status(
    change_id: str,
    current_user: CurrentUser,
    db: DbDep,
):
    """Small status endpoint so the UI can show 'v3 uploaded by X at T' or
    'engine-generated'. Reads only the current latest job.

    Returns 200 with `{version, source, uploaded_at?, reverted_at?,
    supersedes_job?, has_xlsx}`, or 404 when no cert-testcase pack exists
    for this change."""
    _access_check(db, change_id, current_user)

    latest = _latest_succeeded_cert_job(db, change_id)
    if latest is None:
        raise HTTPException(status_code=404, detail="No cert_test_cases pack yet")

    rp = latest.result_payload or {}
    files = rp.get("files") or {}
    xlsx_path = files.get("xlsx") or rp.get("xlsx_path") or ""
    return {
        "job_id":          latest.id,
        "version":         rp.get("version") or 1,
        "source":          rp.get("source") or "engine",
        "uploaded_at":     rp.get("uploaded_at"),
        "reverted_at":     rp.get("reverted_at"),
        "supersedes_job":  rp.get("supersedes_job"),
        "reverted_to_job": rp.get("reverted_to_job"),
        "has_xlsx":        bool(xlsx_path) and Path(xlsx_path).exists(),
        "warnings":        rp.get("warnings") or [],
        "original_filename": rp.get("original_filename") or "",
    }


__all__ = ["router"]
